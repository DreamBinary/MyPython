# workbool工作簿 worksheet表单 行、列、单元格（row column cell）


import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
a=openpyxl.load_workbook('我好帅.xlsx')
print(a.sheetnames)
for sheet in a:
    print(sheet.title)
mysheet=a.create_sheet('啊')
print(a.sheetnames)
#sheet1=a.get_sheet_by_name('我')
b=a.active
print(b)
print(b['A1'])
print(b['A1'].value)
c=b['A2']
print('Row {},Column {} is {}'.format(c.row,c.column,c.value))
print('Cell {} is {}'.format(c.coordinate,c.value)) #坐标 值
print('--------------------------------')
# print(b.cell(row=3,column=A).value) 报错
print(b.cell(row=3,column=1).value)
# print(b.cell[coordinate=A3].value)
# print(b.cell['A1'].value)
for i in range(1,7):
    print(i, b.cell(row=i, column=1).value)
colC=b['C']
print(colC)
print(colC[2].value)
row3=b[3]
print(row3)
col_range=b['A:B']
print(col_range)
print('------------------------')
# for cell in col_range:
#     print(cell.value)  错
for col in col_range:
    print(col)
    # print(col.value)
print('--------------------')
for col in col_range:
    for cell in col:
        print(cell.value)
print('--------------------')
for row in b.iter_rows(min_row=1,max_row=2,min_col=2,max_col=3):
    for cell in row:
        print(cell.value)
print('--------------------')
print(tuple(b.rows))
print('--------------------')
cell_range=b['A1:C3']
for e_cell in cell_range:
    for cell in e_cell:
        print(cell.coordinate,cell.value)
# for cell in cell_range:   错
#     print(cell.coordinate, cell.value)
print('--------------------')
print('{}*{}'.format(b.max_row,b.max_column))
print("---------------------")
print(get_column_letter(2),get_column_letter(5),get_column_letter(520))
print(column_index_from_string('AAH'))






















