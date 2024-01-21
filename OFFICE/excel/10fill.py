from openpyxl.styles import PatternFill,GradientFill
from openpyxl import load_workbook
wb=load_workbook(filename='排序与筛选.xlsx')
sheet=wb['成绩表']
cell1=sheet['D1']
pattern_fill=PatternFill(fill_type='solid',fgColor='99ccff')
cell1.fill=pattern_fill
cell2=sheet['F1']
gradient_fill=GradientFill(stop=('FFFFFF',"99ccff",'000000'))
cell2.fill=gradient_fill
wb.save(filename='填充样式.xlsx')























