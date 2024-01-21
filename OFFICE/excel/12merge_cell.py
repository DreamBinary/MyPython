import openpyxl
from openpyxl import load_workbook
wb=load_workbook('排序与筛选.xlsx')
sheet=wb['成绩表']
sheet.merge_cells("A3:G3")
sheet.merge_cells(start_row=8,end_row=12,start_column=1,end_column=3)
wb.save('合并单元格.xlsx')

#取消合并单元格   .unmerge_cells
























