import openpyxl

print(dir(openpyxl))
wb = openpyxl.load_workbook(filename='排序与筛选.xlsx')
print('1.0', wb.sheetnames)
print('1.05', wb.get_sheet_names())
# 指定sheet
print('1.1', wb._sheets[1])
print('1.2', wb.active)
print('1.3', wb.get_sheet_by_name('成绩表'))
print('1.4', wb['成绩表'])
print('1.5', wb.worksheets[0], wb.worksheets[1], wb.worksheets[2], wb.worksheets[3])

# sheet=wb['shuai']
# print(sheet.dimensions)
sheet = wb.active  # 打开文件时的sheet
print('2', sheet)
cell = sheet['F5']
print('3', cell.value)
cell1 = sheet.cell(2, 3)
print('4', cell1, cell1.value)
print('5', cell, cell.value, cell.row, cell.column, cell.coordinate)

sheet2 = wb.worksheets[3]
cell2 = sheet2['B3:E6']
print('6', cell2)
# print('7',cell2.value)
for i in cell2:  # 行
    for j in i:  # 列
        print('7', j.value)
# for i in cell2:
#     print('8',i.value)

# 按行、列获取值
for i in sheet2.iter_rows(min_row=2, min_col=1, max_col=2):
    for j in i:
        print("8", j.value)
for i in sheet2.iter_cols(min_col=1, max_col=2):
    for j in i:
        print('9', j.value)
print('10', sheet2.rows)
for i in sheet2.rows:
    print(i)
