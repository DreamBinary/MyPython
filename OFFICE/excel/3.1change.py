from openpyxl import load_workbook
wb=load_workbook(filename='排序与筛选.xlsx')
sheet=wb.get_sheet_by_name('成绩表')
print(sheet)
#创建shseet
wb.create_sheet('chang')
print(wb.sheetnames)
wb.save(filename='创建.xlsx')




























